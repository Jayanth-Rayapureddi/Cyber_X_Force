import os
from io import BytesIO
from datetime import date, datetime, time
from typing import Any

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


API_BASE_URL = os.getenv("API_BASE_URL", "http://backend:8000").rstrip("/")
REQUEST_TIMEOUT = 12

st.set_page_config(
    page_title="Cyber_X_Force",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    .block-container {
        padding-top: 1.4rem;
        padding-bottom: 2rem;
    }
    [data-testid="stSidebar"] {
        border-right: 1px solid rgba(255,255,255,0.08);
    }
    .hero {
        padding: 1.2rem 1.35rem;
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 18px;
        background: linear-gradient(135deg, rgba(35,92,171,0.28), rgba(20,25,38,0.45));
        margin-bottom: 1.2rem;
    }
    .hero h1 {
        margin: 0;
        font-size: 2.1rem;
    }
    .hero p {
        margin: 0.35rem 0 0 0;
        opacity: 0.75;
    }
    .kpi-card {
        padding: 1rem 1.05rem;
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 16px;
        background: rgba(255,255,255,0.025);
        min-height: 112px;
    }
    .kpi-label {
        opacity: 0.72;
        font-size: 0.78rem;
        text-transform: uppercase;
        letter-spacing: 0.06em;
    }
    .kpi-value {
        font-size: 1.85rem;
        font-weight: 750;
        margin-top: 0.25rem;
    }
    .kpi-note {
        opacity: 0.62;
        font-size: 0.78rem;
        margin-top: 0.2rem;
    }
    .section-card {
        padding: 1rem;
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 16px;
        background: rgba(255,255,255,0.018);
        margin-bottom: 0.75rem;
    }
    .status-badge {
        display: inline-block;
        padding: 0.22rem 0.55rem;
        border-radius: 999px;
        font-size: 0.75rem;
        font-weight: 700;
        background: rgba(70,130,220,0.18);
        border: 1px solid rgba(120,170,255,0.28);
    }
    .danger-badge {
        display: inline-block;
        padding: 0.22rem 0.55rem;
        border-radius: 999px;
        font-size: 0.75rem;
        font-weight: 700;
        background: rgba(225,70,70,0.16);
        border: 1px solid rgba(255,105,105,0.3);
    }
    .warn-badge {
        display: inline-block;
        padding: 0.22rem 0.55rem;
        border-radius: 999px;
        font-size: 0.75rem;
        font-weight: 700;
        background: rgba(230,170,50,0.16);
        border: 1px solid rgba(255,205,90,0.3);
    }
    .success-badge {
        display: inline-block;
        padding: 0.22rem 0.55rem;
        border-radius: 999px;
        font-size: 0.75rem;
        font-weight: 700;
        background: rgba(50,190,125,0.16);
        border: 1px solid rgba(90,230,160,0.3);
    }
    div[data-testid="stDataFrame"] {
        border: 1px solid rgba(255,255,255,0.07);
        border-radius: 12px;
        overflow: hidden;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def dt_value(day: date, clock: time = time(0, 0)) -> str:
    return datetime.combine(day, clock).isoformat()


def api_request(
    method: str,
    path: str,
    payload: dict[str, Any] | None = None,
) -> Any:
    try:
        headers = {}
        token = st.session_state.get("access_token")
        if token:
            headers["Authorization"] = f"Bearer {token}"
        response = requests.request(
            method,
            f"{API_BASE_URL}{path}",
            json=payload,
            headers=headers,
            timeout=REQUEST_TIMEOUT,
        )
        if response.status_code >= 400:
            detail: Any = response.text
            try:
                detail = response.json().get("detail", detail)
            except ValueError:
                pass
            st.error(f"{method} {path} failed: {detail}")
            return None
        if response.status_code == 204:
            return True
        return response.json()
    except requests.RequestException as exc:
        st.error(f"Backend request failed for {path}: {exc}")
        return None


def api_get(path: str) -> Any:
    return api_request("GET", path)


def api_post(path: str, payload: dict[str, Any]) -> Any:
    return api_request("POST", path, payload)


def api_put(path: str, payload: dict[str, Any]) -> Any:
    return api_request("PUT", path, payload)


def api_delete(path: str) -> bool:
    return bool(api_request("DELETE", path))


def api_upload_evidence(data: dict[str, Any], uploaded_file) -> Any:
    try:
        response = requests.post(
            f"{API_BASE_URL}/evidence/upload",
            data=data,
            files={"file": (uploaded_file.name, uploaded_file.getvalue(), uploaded_file.type)},
            headers={"Authorization": f"Bearer {st.session_state.get('access_token', '')}"},
            timeout=60,
        )
        if response.status_code >= 400:
            detail: Any = response.text
            try:
                detail = response.json().get("detail", detail)
            except ValueError:
                pass
            st.error(f"Evidence upload failed: {detail}")
            return None
        return response.json()
    except requests.RequestException as exc:
        st.error(f"Evidence upload failed: {exc}")
        return None


def download_evidence_file(evidence_id: int) -> tuple[bytes, str] | None:
    try:
        response = requests.get(
            f"{API_BASE_URL}/evidence/{evidence_id}/download",
            headers={"Authorization": f"Bearer {st.session_state.get('access_token', '')}"},
            timeout=60,
        )
        response.raise_for_status()
        disposition = response.headers.get("content-disposition", "")
        filename = f"evidence-{evidence_id}"
        if "filename=" in disposition:
            filename = disposition.split("filename=", 1)[1].strip().strip('"')
        return response.content, filename
    except requests.RequestException as exc:
        st.error(f"Download failed: {exc}")
        return None



def readable_column_name(name: str) -> str:
    return name.replace("_", " ").strip().title()


def excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        return str(value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def build_excel_workbook(
    sheets: dict[str, list[dict[str, Any]]],
    title: str,
) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True)
    generated_font = Font(size=10, italic=True, color="666666")

    for sheet_name, records in sheets.items():
        safe_name = sheet_name[:31] or "Data"
        worksheet = workbook.create_sheet(title=safe_name)

        worksheet["A1"] = title
        worksheet["A1"].font = title_font
        worksheet["A2"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"
        worksheet["A2"].font = generated_font

        if not records:
            worksheet["A4"] = "No records available."
            worksheet.column_dimensions["A"].width = 28
            continue

        frame = pd.DataFrame(records)
        frame = frame.drop(
            columns=[
                column
                for column in ["hashed_password"]
                if column in frame.columns
            ],
            errors="ignore",
        )

        headers = list(frame.columns)
        header_row = 4

        for column_index, column_name in enumerate(headers, start=1):
            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=readable_column_name(column_name),
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index, row in enumerate(
            frame.itertuples(index=False, name=None),
            start=header_row + 1,
        ):
            for column_index, value in enumerate(row, start=1):
                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=excel_value(value),
                )
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_index, column_name in enumerate(headers, start=1):
            values = [
                readable_column_name(column_name),
                *[
                    str(excel_value(value))
                    for value in frame.iloc[:, column_index - 1].tolist()
                ],
            ]
            width = min(max(len(value) for value in values) + 2, 45)
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = max(width, 12)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def excel_download_button(
    label: str,
    file_name: str,
    sheets: dict[str, list[dict[str, Any]]],
    title: str,
    key: str,
) -> None:
    data = build_excel_workbook(sheets, title)
    st.download_button(
        label,
        data=data,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
        width="stretch",
    )



PDF_REPORT_COLUMNS: dict[str, list[str]] = {
    "Risk Register": [
        "risk_code", "title", "inherent_level", "inherent_score",
        "likelihood", "impact", "status", "treatment_option", "risk_owner",
        "residual_level",
    ],
    "ISO Controls": [
        "control_code", "title", "category", "implementation_status",
        "implementation_percentage", "owner", "target_date",
    ],
    "Compliance Assessments": [
        "assessment_code", "control_id", "compliance_status",
        "compliance_score", "assessor", "assessment_date", "next_review_date",
    ],
    "Audits": [
        "audit_code", "title", "audit_type", "lead_auditor",
        "planned_start_date", "planned_end_date", "status",
    ],
    "Audit Findings": [
        "finding_code", "title", "finding_type", "severity", "status",
        "owner", "due_date", "audit_id", "control_id",
    ],
    "Corrective Actions": [
        "action_code", "action_description", "action_owner", "status",
        "completion_percentage", "target_date", "completion_date", "finding_id",
    ],
    "Incidents": [
        "incident_code", "title", "category", "severity", "status",
        "assigned_to", "reported_by", "detected_at", "asset_id",
    ],
    "Evidence Register": [
        "evidence_code", "title", "evidence_type", "owner", "control_id",
        "assessment_id", "collected_at", "valid_until", "is_verified",
    ],
}


def pdf_safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (dict, list, tuple, set)):
        return str(value)
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def report_styles() -> dict[str, Any]:
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="ReportTitle",
            parent=styles["Title"],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#17365D"),
            alignment=TA_CENTER,
            spaceAfter=12,
        )
    )
    styles.add(
        ParagraphStyle(
            name="ReportSubtitle",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#5A6470"),
            alignment=TA_CENTER,
            spaceAfter=16,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionHeading",
            parent=styles["Heading2"],
            fontSize=14,
            leading=17,
            textColor=colors.HexColor("#1F4E78"),
            spaceBefore=8,
            spaceAfter=8,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SmallTable",
            parent=styles["BodyText"],
            fontSize=7,
            leading=9,
        )
    )
    return styles


def add_page_number(canvas, document) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    page_text = f"Cyber_X_Force | Page {document.page}"
    canvas.drawRightString(285 * mm, 10 * mm, page_text)
    canvas.restoreState()


def summary_table(summary_items: list[tuple[str, Any]], styles: dict[str, Any]) -> Table:
    cells = []
    for label, value in summary_items:
        cells.append(
            [
                Paragraph(f"<b>{pdf_safe_text(label)}</b>", styles["SmallTable"]),
                Paragraph(pdf_safe_text(value), styles["SmallTable"]),
            ]
        )
    table = Table(cells, colWidths=[55 * mm, 65 * mm], hAlign="CENTER")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D9EAF7")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B5C7D8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def records_table(
    section_name: str,
    records: list[dict[str, Any]],
    styles: dict[str, Any],
) -> Table | Paragraph:
    if not records:
        return Paragraph("No records available.", styles["BodyText"])

    preferred = PDF_REPORT_COLUMNS.get(section_name, [])
    available = list(records[0].keys())
    columns = [column for column in preferred if column in available]
    if not columns:
        columns = [
            column for column in available
            if column not in {"created_at", "updated_at", "hashed_password"}
        ][:10]

    header = [
        Paragraph(f"<b>{readable_column_name(column)}</b>", styles["SmallTable"])
        for column in columns
    ]
    rows = [header]
    for record in records:
        rows.append(
            [Paragraph(pdf_safe_text(record.get(column)), styles["SmallTable"]) for column in columns]
        )

    page_width = landscape(A4)[0] - 24 * mm
    col_width = page_width / max(len(columns), 1)
    table = Table(rows, colWidths=[col_width] * len(columns), repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C4CE")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F9")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def build_pdf_report(
    title: str,
    subtitle: str,
    summary_items: list[tuple[str, Any]],
    sections: list[tuple[str, list[dict[str, Any]]]],
    generated_by: str,
) -> bytes:
    output = BytesIO()
    styles = report_styles()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=16 * mm,
        title=title,
        author="Cyber_X_Force",
        subject=subtitle,
    )

    story: list[Any] = [
        Paragraph(pdf_safe_text(title), styles["ReportTitle"]),
        Paragraph(
            f"{pdf_safe_text(subtitle)}<br/>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Generated by: {pdf_safe_text(generated_by)}",
            styles["ReportSubtitle"],
        ),
    ]

    if summary_items:
        story.append(summary_table(summary_items, styles))
        story.append(Spacer(1, 8 * mm))

    for index, (section_name, records) in enumerate(sections):
        if index > 0:
            story.append(PageBreak())
        story.append(Paragraph(pdf_safe_text(section_name), styles["SectionHeading"]))
        story.append(records_table(section_name, records, styles))

    document.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    return output.getvalue()


def clean_frame(records: list[dict[str, Any]], preferred: list[str]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    frame = pd.DataFrame(records)
    visible = [column for column in preferred if column in frame.columns]
    remaining = [
        column
        for column in frame.columns
        if column not in visible
        and column not in {"created_at", "updated_at", "hashed_password"}
    ]
    return frame[visible + remaining]


def show_table(
    records: list[dict[str, Any]],
    preferred: list[str],
    empty_message: str,
) -> None:
    frame = clean_frame(records, preferred)
    if frame.empty:
        st.info(empty_message)
        return
    st.dataframe(frame, width="stretch", hide_index=True)


def kpi(label: str, value: Any, note: str = "") -> None:
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div>
            <div class="kpi-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def badge_class(value: str) -> str:
    lowered = value.lower()
    if any(word in lowered for word in ["critical", "non-compliant", "overdue", "major"]):
        return "danger-badge"
    if any(word in lowered for word in ["high", "planned", "open", "in progress", "partial"]):
        return "warn-badge"
    if any(word in lowered for word in ["implemented", "compliant", "completed", "closed", "verified"]):
        return "success-badge"
    return "status-badge"


def page_header(title: str, subtitle: str) -> None:
    st.markdown(
        f"""
        <div class="hero">
            <h1>{title}</h1>
            <p>{subtitle}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def dashboard_page() -> None:
    page_header(
        "🛡️ Cyber_X_Force",
        "Executive view of the ISO/IEC 27001 risk, compliance, audit and incident posture.",
    )
    management = api_get("/management/dashboard")
    compliance = api_get("/compliance/dashboard")
    risks = api_get("/risks") or []

    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export risk register",
            "cyber_x_force_risk_register.xlsx",
            {"Risk Register": risks},
            "Cyber_X_Force Risk Register",
            "export-risks",
        )
    incidents = api_get("/incidents") or []
    findings = api_get("/audit-findings") or []
    actions = api_get("/corrective-actions") or []
    evidence = api_get("/evidence") or []
    audits = api_get("/audits") or []

    if not management:
        return

    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export management workbook",
            "cyber_x_force_management_dashboard.xlsx",
            {
                "Management Summary": [management],
                "Compliance Dashboard": [compliance or {}],
                "Risks": risks,
                "Incidents": incidents,
                "Audit Findings": findings,
            },
            "Cyber_X_Force Management Dashboard",
            "export-management",
        )

    row1 = st.columns(4)
    with row1[0]:
        kpi("Total Assets", management.get("total_assets", 0), "Registered information assets")
    with row1[1]:
        kpi("Total Risks", management.get("total_risks", 0), "Risks in the register")
    with row1[2]:
        kpi("Critical Risks", management.get("critical_risks", 0), "Immediate management attention")
    with row1[3]:
        kpi("Compliance", f"{management.get('compliance_percentage', 0):.2f}%", "Applicable controls implemented")

    row2 = st.columns(4)
    with row2[0]:
        kpi("Open Incidents", management.get("open_incidents", 0), "Not yet closed")
    with row2[1]:
        kpi("Open Findings", management.get("open_audit_findings", 0), "Audit findings requiring action")
    with row2[2]:
        kpi("Overdue Actions", management.get("overdue_corrective_actions", 0), "Corrective actions past due")
    with row2[3]:
        kpi("API Status", "Connected", API_BASE_URL)

    st.write("")
    left, right = st.columns(2)

    with left:
        st.subheader("Risk profile")
        if risks:
            risk_df = pd.DataFrame(risks)
            counts = (
                risk_df["inherent_level"]
                .value_counts()
                .rename_axis("level")
                .reset_index(name="count")
            )
            fig = px.bar(
                counts,
                x="level",
                y="count",
                text_auto=True,
                category_orders={"level": ["Low", "Medium", "High", "Critical"]},
            )
            fig.update_layout(height=340, margin=dict(l=10, r=10, t=20, b=10))
            st.plotly_chart(fig, width="stretch")
        else:
            st.info("No risk data is available.")

    with right:
        st.subheader("Compliance readiness")
        percentage = float(management.get("compliance_percentage", 0))
        gauge = go.Figure(
            go.Indicator(
                mode="gauge+number",
                value=percentage,
                number={"suffix": "%"},
                gauge={"axis": {"range": [0, 100]}},
            )
        )
        gauge.update_layout(height=340, margin=dict(l=25, r=25, t=15, b=10))
        st.plotly_chart(gauge, width="stretch")

    left, right = st.columns(2)
    with left:
        st.subheader("Control status")
        status_data = (compliance or {}).get("status_distribution", [])
        if status_data:
            fig = px.pie(
                pd.DataFrame(status_data),
                names="status",
                values="count",
                hole=0.52,
            )
            fig.update_layout(height=340, margin=dict(l=10, r=10, t=20, b=10))
            st.plotly_chart(fig, width="stretch")
        else:
            st.info("No control-status data is available.")

    with right:
        st.subheader("Current operational workload")
        workload = pd.DataFrame(
            [
                {"area": "Open incidents", "count": management.get("open_incidents", 0)},
                {"area": "Open findings", "count": management.get("open_audit_findings", 0)},
                {"area": "Overdue actions", "count": management.get("overdue_corrective_actions", 0)},
            ]
        )
        fig = px.bar(workload, x="area", y="count", text_auto=True)
        fig.update_layout(height=340, margin=dict(l=10, r=10, t=20, b=10))
        st.plotly_chart(fig, width="stretch")

    st.subheader("Top overdue controls")
    overdue = (compliance or {}).get("top_overdue_controls", [])
    if not overdue:
        st.success("No overdue controls.")
    else:
        overdue_rows = [
            {
                "Control": control.get("control_code"),
                "Title": control.get("title"),
                "Owner": control.get("owner") or "Unassigned",
                "Status": control.get("implementation_status"),
                "Progress": f"{control.get('implementation_percentage', 0)}%",
                "Target Date": str(control.get("target_date", ""))[:10],
                "Days Overdue": control.get("days_overdue", 0),
            }
            for control in overdue[:5]
        ]
        show_table(
            overdue_rows,
            [
                "Control",
                "Title",
                "Owner",
                "Status",
                "Progress",
                "Target Date",
                "Days Overdue",
            ],
            "No overdue controls.",
        )

    if incidents or findings:
        
        st.subheader("Recent governance activity")

        activity_left, activity_right = st.columns(2)

        with activity_left:
            st.markdown("### Latest incidents")

            incident_rows = [
                {
                    "Incident Code": item.get("incident_code"),
                    "Title": item.get("title"),
                    "Severity": item.get("severity"),
                    "Status": item.get("status"),
                    "Assigned To": item.get("assigned_to"),
                    "Detected Date": str(item.get("detected_at", ""))[:10],
                }
                for item in incidents[:5]
            ]

            show_table(
                incident_rows,
                [
                    "Incident Code",
                    "Title",
                    "Severity",
                    "Status",
                    "Assigned To",
                    "Detected Date",
                ],
                "No incidents.",
            )

        with activity_right:
            st.markdown("### Latest findings")

            finding_rows = [
                {
                    "Finding Code": item.get("finding_code"),
                    "Title": item.get("title"),
                    "Severity": item.get("severity"),
                    "Owner": item.get("owner"),
                    "Status": item.get("status"),
                    "Due Date": str(item.get("due_date", ""))[:10],
                }
                for item in findings[:5]
            ]

            show_table(
                finding_rows,
                [
                    "Finding Code",
                    "Title",
                    "Severity",
                    "Owner",
                    "Status",
                    "Due Date",
                ],
                "No findings.",
            )


    st.subheader("Executive priorities")

    priority_left, priority_middle, priority_right = st.columns(3)

    with priority_left:
        st.markdown("### Top critical risks")

        critical_risks = [
            risk
            for risk in risks
            if risk.get("inherent_level") in ["Critical", "High"]
        ]

        critical_risks = sorted(
            critical_risks,
            key=lambda item: int(item.get("inherent_score", 0)),
            reverse=True,
        )[:5]

        if not critical_risks:
            st.success("No critical or high risks.")
        else:
            for risk in critical_risks:
                level = str(risk.get("inherent_level", ""))
                st.markdown(
                    f"""
                    <div class="section-card">
                        <span class="{badge_class(level)}">{level}</span>
                        <h4>{risk.get("risk_code", "")}</h4>
                        <b>{risk.get("title", "")}</b><br><br>
                        Score: {risk.get("inherent_score", 0)}<br>
                        Owner: {risk.get("risk_owner") or "Unassigned"}<br>
                        Status: {risk.get("status") or "Unknown"}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

    with priority_middle:
        st.markdown("### Latest incidents")

        latest_incidents = sorted(
            incidents,
            key=lambda item: str(item.get("detected_at", "")),
            reverse=True,
        )[:5]

        if not latest_incidents:
            st.info("No incidents have been recorded.")
        else:
            for incident in latest_incidents:
                severity = str(incident.get("severity", ""))
                st.markdown(
                    f"""
                    <div class="section-card">
                        <span class="{badge_class(severity)}">{severity}</span>
                        <span class="status-badge">{incident.get("status", "")}</span>
                        <h4>{incident.get("incident_code", "")}</h4>
                        <b>{incident.get("title", "")}</b><br><br>
                        Assigned to: {incident.get("assigned_to") or "Unassigned"}<br>
                        Detected: {str(incident.get("detected_at", ""))[:10]}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

    with priority_right:
        st.markdown("### Corrective-action progress")

        action_summary = {
            "Open": 0,
            "In Progress": 0,
            "Completed": 0,
            "Verified": 0,
        }

        for action in actions:
            status = action.get("status")
            if status in action_summary:
                action_summary[status] += 1

        action_frame = pd.DataFrame(
            [
                {"Status": status, "Count": count}
                for status, count in action_summary.items()
            ]
        )

        if action_frame["Count"].sum() == 0:
            st.info("No corrective actions have been recorded.")
        else:
            fig = px.bar(
                action_frame,
                x="Status",
                y="Count",
                text_auto=True,
            )
            fig.update_layout(
                height=330,
                margin=dict(l=10, r=10, t=20, b=10),
            )
            st.plotly_chart(fig, width="stretch")



    st.subheader("Governance monitoring")

    monitor_left, monitor_middle, monitor_right = st.columns(3)

    with monitor_left:
        st.markdown("### Upcoming audits")

        upcoming_audits = [
            audit
            for audit in audits
            if audit.get("status") in ["Planned", "In Progress"]
        ]

        upcoming_audits = sorted(
            upcoming_audits,
            key=lambda item: str(item.get("planned_start_date", "")),
        )[:5]

        if not upcoming_audits:
            st.info("No upcoming audits.")
        else:
            for audit in upcoming_audits:
                st.markdown(
                    f"""
                    <div class="section-card">
                        <span class="status-badge">{audit.get("status", "")}</span>
                        <h4>{audit.get("audit_code", "")}</h4>
                        <b>{audit.get("title", "")}</b><br><br>
                        Type: {audit.get("audit_type", "")}<br>
                        Lead auditor: {audit.get("lead_auditor") or "Unassigned"}<br>
                        Start: {str(audit.get("planned_start_date", ""))[:10]}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

    with monitor_middle:
        st.markdown("### Latest evidence")

        latest_evidence = sorted(
            evidence,
            key=lambda item: str(
                item.get("created_at")
                or item.get("collected_at")
                or ""
            ),
            reverse=True,
        )[:5]

        if not latest_evidence:
            st.info("No evidence has been recorded.")
        else:
            for item in latest_evidence:
                verification = (
                    "Verified"
                    if item.get("is_verified")
                    else "Pending verification"
                )

                st.markdown(
                    f"""
                    <div class="section-card">
                        <span class="{badge_class(verification)}">{verification}</span>
                        <h4>{item.get("evidence_code", "")}</h4>
                        <b>{item.get("title", "")}</b><br><br>
                        Type: {item.get("evidence_type", "")}<br>
                        Owner: {item.get("owner") or "Unassigned"}<br>
                        Collected: {str(item.get("collected_at", ""))[:10]}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

    with monitor_right:
        st.markdown("### Findings by severity")

        if not findings:
            st.info("No audit findings have been recorded.")
        else:
            finding_frame = pd.DataFrame(findings)

            severity_counts = (
                finding_frame["severity"]
                .fillna("Unknown")
                .value_counts()
                .rename_axis("Severity")
                .reset_index(name="Count")
            )

            fig = px.pie(
                severity_counts,
                names="Severity",
                values="Count",
                hole=0.5,
            )

            fig.update_layout(
                height=330,
                margin=dict(l=10, r=10, t=20, b=10),
            )

            st.plotly_chart(fig, width="stretch")
    


def assets_page() -> None:
    page_header("Assets", "Register, classify and review organizational information assets.")
    assets = api_get("/assets") or []
    departments = api_get("/departments") or []

    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export assets to Excel",
            "cyber_x_force_assets.xlsx",
            {"Assets": assets},
            "Cyber_X_Force Asset Register",
            "export-assets",
        )

    search_col, type_col, criticality_col = st.columns([2, 1, 1])
    search = search_col.text_input("Search assets", placeholder="Code, name, type or location")
    asset_types = sorted({item.get("asset_type", "") for item in assets if item.get("asset_type")})
    selected_type = type_col.selectbox("Asset type", ["All"] + asset_types)
    criticalities = ["All", "Critical", "High", "Medium", "Low"]
    selected_criticality = criticality_col.selectbox("Criticality", criticalities)

    filtered = assets
    if search:
        token = search.lower()
        filtered = [
            item for item in filtered
            if token in " ".join(
                str(item.get(key, "")).lower()
                for key in ["asset_code", "name", "asset_type", "location"]
            )
        ]
    if selected_type != "All":
        filtered = [item for item in filtered if item.get("asset_type") == selected_type]
    if selected_criticality != "All":
        filtered = [item for item in filtered if item.get("criticality") == selected_criticality]

    show_table(
        filtered,
        ["asset_code", "name", "asset_type", "criticality", "location", "confidentiality", "integrity", "availability", "is_active"],
        "No matching assets.",
    )

    with st.expander("➕ Register a new asset"):
        department_options = {
            f"{item['name']} (ID {item['id']})": item["id"]
            for item in departments
        }
        with st.form("asset_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            asset_code = c1.text_input("Asset code", value="AST-")
            name = c2.text_input("Asset name")
            asset_type = c1.selectbox(
                "Asset type",
                ["Server", "Database", "Application", "Workstation", "Network", "Cloud Service", "Information", "Other"],
            )
            location = c2.text_input("Location")
            description = st.text_area("Description")
            c1, c2, c3 = st.columns(3)
            confidentiality = c1.slider("Confidentiality", 1, 5, 3)
            integrity = c2.slider("Integrity", 1, 5, 3)
            availability = c3.slider("Availability", 1, 5, 3)
            department_label = st.selectbox(
                "Department",
                list(department_options) if department_options else ["No department available"],
            )
            submitted = st.form_submit_button("Create asset", width="stretch")

        if submitted:
            if not department_options:
                st.error("Create a department before registering an asset.")
            else:
                created = api_post(
                    "/assets",
                    {
                        "asset_code": asset_code,
                        "name": name,
                        "asset_type": asset_type,
                        "description": description or None,
                        "location": location or None,
                        "confidentiality": confidentiality,
                        "integrity": integrity,
                        "availability": availability,
                        "department_id": department_options[department_label],
                        "owner_id": None,
                    },
                )
                if created:
                    st.success("Asset created.")
                    st.rerun()


def risk_heatmap(risks: list[dict[str, Any]]) -> None:
    matrix = [[0 for _ in range(5)] for _ in range(5)]
    for risk in risks:
        likelihood = int(risk.get("likelihood", 0))
        impact = int(risk.get("impact", 0))
        if 1 <= likelihood <= 5 and 1 <= impact <= 5:
            matrix[5 - likelihood][impact - 1] += 1
    fig = px.imshow(
        matrix,
        labels={"x": "Impact", "y": "Likelihood", "color": "Risk count"},
        x=[1, 2, 3, 4, 5],
        y=[5, 4, 3, 2, 1],
        text_auto=True,
        aspect="auto",
    )
    fig.update_layout(height=420, margin=dict(l=10, r=10, t=20, b=10))
    st.plotly_chart(fig, width="stretch")


def risks_page() -> None:
    page_header("Risk Register", "Prioritize inherent and residual cybersecurity risks.")
    risks = api_get("/risks") or []

    search_col, level_col, status_col = st.columns([2, 1, 1])
    search = search_col.text_input("Search risks", placeholder="Risk code, title, owner or treatment")
    level = level_col.selectbox("Risk level", ["All", "Critical", "High", "Medium", "Low"])
    status_value = status_col.selectbox("Status", ["All", "Open", "Under Treatment", "Accepted", "Closed"])

    filtered = risks
    if search:
        token = search.lower()
        filtered = [
            item for item in filtered
            if token in " ".join(
                str(item.get(key, "")).lower()
                for key in ["risk_code", "title", "risk_owner", "treatment_option"]
            )
        ]
    if level != "All":
        filtered = [item for item in filtered if item.get("inherent_level") == level]
    if status_value != "All":
        filtered = [item for item in filtered if item.get("status") == status_value]

    if not filtered:
        st.info("No matching risks.")
    else:
        for risk in filtered:
            badge = badge_class(str(risk.get("inherent_level", "")))
            with st.container():
                st.markdown(
                    f"""
                    <div class="section-card">
                        <span class="{badge}">{risk.get('inherent_level', '')}</span>
                        <span class="status-badge">{risk.get('status', '')}</span>
                        <h3>{risk.get('risk_code', '')} — {risk.get('title', '')}</h3>
                        <b>Score:</b> {risk.get('inherent_score', '')}
                        &nbsp;&nbsp; <b>Likelihood:</b> {risk.get('likelihood', '')}/5
                        &nbsp;&nbsp; <b>Impact:</b> {risk.get('impact', '')}/5
                        &nbsp;&nbsp; <b>Treatment:</b> {risk.get('treatment_option', '')}<br>
                        <span style="opacity:.72">{risk.get('description') or 'No description'}</span>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

    st.subheader("5 × 5 risk heat map")
    risk_heatmap(risks)

    show_table(
        risks,
        ["risk_code", "title", "inherent_level", "inherent_score", "likelihood", "impact", "status", "treatment_option", "risk_owner", "residual_level"],
        "No risks have been recorded.",
    )


def controls_page() -> None:
    page_header("ISO Controls", "Monitor control ownership, status and implementation progress.")
    controls = api_get("/controls") or []

    c1, c2, c3 = st.columns([2, 1, 1])
    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export ISO controls",
            "cyber_x_force_iso_controls.xlsx",
            {"ISO Controls": controls},
            "Cyber_X_Force ISO/IEC 27001 Controls",
            "export-controls",
        )

    c1, c2, c3 = st.columns([2, 1, 1])
    search = c1.text_input("Search controls", placeholder="Control code, title or owner")
    categories = sorted({item.get("category") for item in controls if item.get("category")})
    selected_category = c2.selectbox("Category", ["All"] + categories)
    selected_status = c3.selectbox(
        "Status",
        ["All", "Implemented", "In Progress", "Planned", "Not Started", "Not Applicable"],
    )

    filtered = controls
    if search:
        token = search.lower()
        filtered = [
            item for item in filtered
            if token in " ".join(
                str(item.get(key, "")).lower()
                for key in ["control_code", "title", "owner"]
            )
        ]
    if selected_category != "All":
        filtered = [item for item in filtered if item.get("category") == selected_category]
    if selected_status != "All":
        filtered = [item for item in filtered if item.get("implementation_status") == selected_status]

    if not filtered:
        st.info("No matching controls.")
    else:
        for control in filtered:
            status = str(control.get("implementation_status", ""))
            st.markdown(
                f"""
                <div class="section-card">
                    <span class="{badge_class(status)}">{status}</span>
                    <h3>{control.get('control_code', '')} — {control.get('title', '')}</h3>
                    <b>Owner:</b> {control.get('owner') or 'Unassigned'}
                    &nbsp;&nbsp; <b>Category:</b> {control.get('category', '')}
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.progress(int(control.get("implementation_percentage", 0)) / 100)
            st.caption(f"{control.get('implementation_percentage', 0)}% implemented")

    show_table(
        filtered,
        ["control_code", "title", "category", "implementation_status", "implementation_percentage", "owner", "target_date", "last_reviewed_at"],
        "No controls available.",
    )


def compliance_page() -> None:
    page_header("Compliance", "Assess control conformance and monitor ISO/IEC 27001 readiness.")
    summary = api_get("/compliance/summary") or {}
    overdue = api_get("/compliance/overdue-controls") or []
    assessments = api_get("/compliance-assessments") or []
    controls = api_get("/controls") or []

    cols = st.columns(4)
    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export compliance workbook",
            "cyber_x_force_compliance.xlsx",
            {
                "Summary": [summary],
                "Assessments": assessments,
                "Overdue Controls": overdue,
                "ISO Controls": controls,
            },
            "Cyber_X_Force Compliance Report",
            "export-compliance",
        )

    cols = st.columns(4)
    with cols[0]:
        kpi("Total Controls", summary.get("total_controls", 0), "In the control library")
    with cols[1]:
        kpi("Implemented", summary.get("implemented", 0), "Fully implemented controls")
    with cols[2]:
        kpi("Compliance", f"{summary.get('overall_compliance_percentage', 0):.2f}%", "Excludes not-applicable controls")
    with cols[3]:
        kpi("Average Progress", f"{summary.get('average_implementation_percentage', 0):.2f}%", "Average implementation percentage")

    left, right = st.columns([1, 1])
    with left:
        st.subheader("Compliance gauge")
        value = float(summary.get("overall_compliance_percentage", 0))
        fig = go.Figure(go.Indicator(mode="gauge+number", value=value, number={"suffix": "%"}, gauge={"axis": {"range": [0, 100]}}))
        fig.update_layout(height=320, margin=dict(l=20, r=20, t=20, b=10))
        st.plotly_chart(fig, width="stretch")
    with right:
        st.subheader("Assessment status")
        if assessments:
            frame = pd.DataFrame(assessments)
            distribution = frame["compliance_status"].value_counts().rename_axis("status").reset_index(name="count")
            fig = px.pie(distribution, names="status", values="count", hole=0.5)
            fig.update_layout(height=320, margin=dict(l=20, r=20, t=20, b=10))
            st.plotly_chart(fig, width="stretch")
        else:
            st.info("No assessments have been recorded.")

    st.subheader("Overdue controls")
    show_table(
        overdue,
        ["control_code", "title", "owner", "implementation_status", "implementation_percentage", "target_date", "days_overdue"],
        "No overdue controls.",
    )

    with st.expander("➕ Record a compliance assessment"):
        control_options = {
            f"{item['control_code']} — {item['title']}": item["id"]
            for item in controls
        }
        with st.form("assessment_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            assessment_code = c1.text_input("Assessment code", value="ASM-")
            control_label = c2.selectbox("ISO control", list(control_options) if control_options else ["No controls"])
            compliance_status = c1.selectbox(
                "Compliance status",
                ["Not Assessed", "Compliant", "Partially Compliant", "Non-Compliant", "Not Applicable"],
            )
            compliance_score = c2.slider("Compliance score", 0, 100, 50)
            assessor = c1.text_input("Assessor")
            assessment_date = c2.date_input("Assessment date", value=date.today())
            findings = st.text_area("Findings")
            recommendations = st.text_area("Recommendations")
            evidence_reference = st.text_input("Evidence reference")
            next_review = st.date_input("Next review date", value=date.today())
            submitted = st.form_submit_button("Create assessment", width="stretch")
        if submitted and control_options:
            created = api_post(
                "/compliance-assessments",
                {
                    "assessment_code": assessment_code,
                    "control_id": control_options[control_label],
                    "compliance_status": compliance_status,
                    "compliance_score": compliance_score,
                    "assessor": assessor,
                    "assessment_date": dt_value(assessment_date),
                    "findings": findings or None,
                    "recommendations": recommendations or None,
                    "evidence_reference": evidence_reference or None,
                    "next_review_date": dt_value(next_review),
                },
            )
            if created:
                st.success("Compliance assessment created.")
                st.rerun()

    st.subheader("Assessment history")
    show_table(
        assessments,
        ["assessment_code", "control_id", "compliance_status", "compliance_score", "assessor", "assessment_date", "next_review_date"],
        "No compliance assessments.",
    )


def evidence_page() -> None:
    page_header("Evidence Repository", "Upload, verify, download and manage ISO/IEC 27001 audit evidence.")
    evidence = api_get("/evidence") or []
    controls = api_get("/controls") or []
    assessments = api_get("/compliance-assessments") or []

    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export evidence register",
            "cyber_x_force_evidence_register.xlsx",
            {"Evidence Register": evidence},
            "Cyber_X_Force Evidence Register",
            "export-evidence",
        )

    verified = sum(bool(item.get("is_verified")) for item in evidence)
    cols = st.columns(3)
    with cols[0]: kpi("Evidence Records", len(evidence), "Uploaded and referenced evidence")
    with cols[1]: kpi("Verified", verified, "Reviewed by an auditor")
    with cols[2]: kpi("Pending Verification", len(evidence) - verified, "Awaiting review")

    with st.expander("➕ Upload evidence", expanded=not evidence):
        control_options = {"None": ""} | {f"{item['control_code']} — {item['title']}": str(item["id"]) for item in controls}
        assessment_options = {"None": ""} | {item["assessment_code"]: str(item["id"]) for item in assessments}
        with st.form("evidence_upload_form", clear_on_submit=True):
            uploaded_file = st.file_uploader("Choose evidence file", type=["pdf", "docx", "xlsx", "xls", "png", "jpg", "jpeg", "txt", "csv"], help="Maximum size: 10 MB")
            c1, c2 = st.columns(2)
            evidence_code = c1.text_input("Evidence code", value="EVD-")
            title = c2.text_input("Title")
            evidence_type = c1.selectbox("Evidence type", ["Policy", "Procedure", "Report", "Screenshot", "Log", "Certificate", "Spreadsheet", "Other"])
            owner = c2.text_input("Owner")
            description = st.text_area("Description")
            c1, c2 = st.columns(2)
            control_label = c1.selectbox("Related control", list(control_options))
            assessment_label = c2.selectbox("Related assessment", list(assessment_options))
            collected_at = c1.date_input("Collected date", value=date.today())
            has_expiry = c2.checkbox("Evidence has an expiry date")
            valid_until = c2.date_input("Valid until", value=date.today(), disabled=not has_expiry)
            is_verified = st.checkbox("Verified")
            verification_notes = st.text_area("Verification notes")
            submitted = st.form_submit_button("Upload evidence", width="stretch")
        if submitted:
            if uploaded_file is None:
                st.error("Select a file before uploading.")
            elif evidence_code.strip() == "EVD-":
                st.error("Enter a complete evidence code, for example EVD-001.")
            elif not title.strip():
                st.error("Enter an evidence title.")
            else:
                created = api_upload_evidence({
                    "evidence_code": evidence_code.strip(), "title": title.strip(), "evidence_type": evidence_type,
                    "description": description or "", "control_id": control_options[control_label],
                    "assessment_id": assessment_options[assessment_label], "owner": owner or "",
                    "collected_at": dt_value(collected_at), "valid_until": dt_value(valid_until) if has_expiry else "",
                    "is_verified": str(is_verified).lower(), "verification_notes": verification_notes or "",
                }, uploaded_file)
                if created:
                    st.success("Evidence uploaded successfully.")
                    st.rerun()

    search = st.text_input("Search evidence", placeholder="Code, title, type or owner")
    filtered = evidence
    if search:
        token = search.lower()
        filtered = [item for item in evidence if token in " ".join(str(item.get(key, "")).lower() for key in ["evidence_code", "title", "evidence_type", "owner"])]
    if not filtered:
        st.info("No evidence records match your search.")
        return

    for item in filtered:
        left, middle, right = st.columns([5, 1.3, 1.1])
        with left:
            status_label = "Verified" if item.get("is_verified") else "Pending verification"
            st.markdown(f"""<div class='section-card'><span class='{badge_class(status_label)}'>{status_label}</span><h3>{item.get('evidence_code', '')} — {item.get('title', '')}</h3><b>Type:</b> {item.get('evidence_type', '')} &nbsp;&nbsp; <b>Owner:</b> {item.get('owner') or 'Unassigned'}<br><span style='opacity:.72'>{item.get('description') or 'No description'}</span></div>""", unsafe_allow_html=True)
        with middle:
            downloaded = download_evidence_file(item["id"])
            if downloaded:
                content, filename = downloaded
                st.download_button("⬇ Download", data=content, file_name=filename, key=f"download-{item['id']}", width="stretch")
        with right:
            if st.button("🗑 Delete", key=f"delete-{item['id']}", width="stretch"):
                if api_delete(f"/evidence/{item['id']}"):
                    st.success("Evidence deleted.")
                    st.rerun()

    show_table(filtered, ["evidence_code", "title", "evidence_type", "owner", "control_id", "assessment_id", "collected_at", "valid_until", "is_verified"], "No evidence records have been added.")

def audits_page() -> None:
    page_header("Audits and Findings", "Plan internal audits and track non-conformities.")
    audits = api_get("/audits") or []
    findings = api_get("/audit-findings") or []
    controls = api_get("/controls") or []

    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export audits and findings",
            "cyber_x_force_audits_findings.xlsx",
            {
                "Audits": audits,
                "Audit Findings": findings,
                "ISO Controls": controls,
            },
            "Cyber_X_Force Audit and Findings Report",
            "export-audits-findings",
        )

    tab1, tab2 = st.tabs(["Audits", "Audit Findings"])

    with tab1:
        with st.expander("➕ Create an audit"):
            with st.form("audit_form", clear_on_submit=True):
                c1, c2 = st.columns(2)
                audit_code = c1.text_input("Audit code", value="AUD-")
                title = c2.text_input("Audit title")
                audit_type = c1.selectbox("Audit type", ["Internal", "Supplier", "Certification", "Follow-up"])
                lead_auditor = c2.text_input("Lead auditor")
                scope = st.text_area("Scope")
                c1, c2 = st.columns(2)
                start_date = c1.date_input("Planned start", value=date.today())
                end_date = c2.date_input("Planned end", value=date.today())
                status_value = st.selectbox("Status", ["Planned", "In Progress", "Completed", "Cancelled"])
                summary = st.text_area("Summary")
                submitted = st.form_submit_button("Create audit", width="stretch")
            if submitted:
                created = api_post(
                    "/audits",
                    {
                        "audit_code": audit_code,
                        "title": title,
                        "audit_type": audit_type,
                        "scope": scope,
                        "lead_auditor": lead_auditor,
                        "planned_start_date": dt_value(start_date),
                        "planned_end_date": dt_value(end_date),
                        "status": status_value,
                        "summary": summary or None,
                    },
                )
                if created:
                    st.success("Audit created.")
                    st.rerun()

        show_table(
            audits,
            ["audit_code", "title", "audit_type", "lead_auditor", "planned_start_date", "planned_end_date", "status"],
            "No audits have been created.",
        )

    with tab2:
        audit_options = {f"{item['audit_code']} — {item['title']}": item["id"] for item in audits}
        control_options = {"None": None} | {
            f"{item['control_code']} — {item['title']}": item["id"]
            for item in controls
        }
        with st.expander("➕ Record an audit finding"):
            with st.form("finding_form", clear_on_submit=True):
                c1, c2 = st.columns(2)
                finding_code = c1.text_input("Finding code", value="FND-")
                audit_label = c2.selectbox("Audit", list(audit_options) if audit_options else ["No audits"])
                title = c1.text_input("Finding title")
                control_label = c2.selectbox("Related control", list(control_options))
                finding_type = c1.selectbox(
                    "Finding type",
                    ["Observation", "Opportunity for Improvement", "Minor Non-Conformity", "Major Non-Conformity"],
                )
                severity = c2.selectbox("Severity", ["Low", "Medium", "High", "Critical"])
                description = st.text_area("Description")
                root_cause = st.text_area("Root cause")
                owner = c1.text_input("Owner")
                due_date = c2.date_input("Due date", value=date.today())
                status_value = st.selectbox("Status", ["Open", "Under Review", "Remediation", "Verified", "Closed"])
                submitted = st.form_submit_button("Create finding", width="stretch")
            if submitted and audit_options:
                created = api_post(
                    "/audit-findings",
                    {
                        "finding_code": finding_code,
                        "audit_id": audit_options[audit_label],
                        "control_id": control_options[control_label],
                        "title": title,
                        "finding_type": finding_type,
                        "severity": severity,
                        "description": description,
                        "root_cause": root_cause or None,
                        "owner": owner or None,
                        "due_date": dt_value(due_date),
                        "status": status_value,
                    },
                )
                if created:
                    st.success("Audit finding created.")
                    st.rerun()

        show_table(
            findings,
            ["finding_code", "title", "finding_type", "severity", "status", "owner", "due_date", "audit_id", "control_id"],
            "No audit findings.",
        )


def actions_page() -> None:
    page_header("Corrective Actions", "Track remediation from open action through verification.")
    actions = api_get("/corrective-actions") or []
    findings = api_get("/audit-findings") or []

    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export corrective actions",
            "cyber_x_force_corrective_actions.xlsx",
            {
                "Corrective Actions": actions,
                "Related Findings": findings,
            },
            "Cyber_X_Force Corrective Action Register",
            "export-actions",
        )

    finding_options = {
        f"{item['finding_code']} — {item['title']}": item["id"]
        for item in findings
    }
    with st.expander("➕ Create a corrective action"):
        with st.form("corrective_action_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            action_code = c1.text_input("Action code", value="CAP-")
            finding_label = c2.selectbox("Audit finding", list(finding_options) if finding_options else ["No findings"])
            owner = c1.text_input("Action owner")
            target_date = c2.date_input("Target date", value=date.today())
            description = st.text_area("Action description")
            status_value = c1.selectbox("Status", ["Open", "In Progress", "Completed", "Verified", "Cancelled"])
            completion = c2.slider("Completion percentage", 0, 100, 0)
            verification = st.text_area("Verification result")
            submitted = st.form_submit_button("Create corrective action", width="stretch")
        if submitted and finding_options:
            created = api_post(
                "/corrective-actions",
                {
                    "action_code": action_code,
                    "finding_id": finding_options[finding_label],
                    "action_description": description,
                    "action_owner": owner,
                    "target_date": dt_value(target_date),
                    "status": status_value,
                    "completion_percentage": completion,
                    "completion_date": dt_value(date.today()) if completion == 100 else None,
                    "verification_result": verification or None,
                },
            )
            if created:
                st.success("Corrective action created.")
                st.rerun()

    statuses = ["Open", "In Progress", "Completed", "Verified"]
    columns = st.columns(4)
    for index, status_value in enumerate(statuses):
        with columns[index]:
            st.subheader(status_value)
            matching = [item for item in actions if item.get("status") == status_value]
            if not matching:
                st.caption("No actions")
            for item in matching:
                st.markdown(
                    f"""
                    <div class="section-card">
                        <b>{item.get('action_code', '')}</b><br>
                        {item.get('action_description', '')}<br><br>
                        Owner: {item.get('action_owner', '')}<br>
                        Due: {str(item.get('target_date', ''))[:10]}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                st.progress(int(item.get("completion_percentage", 0)) / 100)

    show_table(
        actions,
        ["action_code", "action_description", "action_owner", "status", "completion_percentage", "target_date", "completion_date", "finding_id"],
        "No corrective actions.",
    )


def incidents_page() -> None:
    page_header("Incidents", "Record, triage, contain and close cybersecurity incidents.")
    incidents = api_get("/incidents") or []
    assets = api_get("/assets") or []

    export_col, _ = st.columns([1, 3])
    with export_col:
        excel_download_button(
            "⬇ Export incident register",
            "cyber_x_force_incidents.xlsx",
            {
                "Incidents": incidents,
                "Assets": assets,
            },
            "Cyber_X_Force Incident Register",
            "export-incidents",
        )

    asset_options = {"None": None} | {
        f"{item['asset_code']} — {item['name']}": item["id"]
        for item in assets
    }

    with st.expander("➕ Report an incident"):
        with st.form("incident_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            incident_code = c1.text_input("Incident code", value="INC-")
            title = c2.text_input("Title")
            category = c1.selectbox(
                "Category",
                ["Malware", "Phishing", "Unauthorized Access", "Data Breach", "Service Disruption", "Insider Threat", "Other"],
            )
            severity = c2.selectbox("Severity", ["Low", "Medium", "High", "Critical"])
            description = st.text_area("Description")
            asset_label = c1.selectbox("Affected asset", list(asset_options))
            reported_by = c2.text_input("Reported by")
            assigned_to = c1.text_input("Assigned to")
            detected_date = c2.date_input("Detected date", value=date.today())
            status_value = st.selectbox(
                "Status",
                ["Reported", "Triaged", "Investigating", "Contained", "Recovered", "Closed"],
            )
            submitted = st.form_submit_button("Create incident", width="stretch")
        if submitted:
            created = api_post(
                "/incidents",
                {
                    "incident_code": incident_code,
                    "title": title,
                    "category": category,
                    "severity": severity,
                    "description": description,
                    "asset_id": asset_options[asset_label],
                    "reported_by": reported_by,
                    "assigned_to": assigned_to or None,
                    "detected_at": dt_value(detected_date),
                    "status": status_value,
                    "containment_summary": None,
                    "root_cause": None,
                    "lessons_learned": None,
                    "closed_at": dt_value(date.today()) if status_value == "Closed" else None,
                },
            )
            if created:
                st.success("Incident created.")
                st.rerun()

    severity_filter = st.multiselect(
        "Severity filter",
        ["Critical", "High", "Medium", "Low"],
        default=["Critical", "High", "Medium", "Low"],
    )
    filtered = [item for item in incidents if item.get("severity") in severity_filter]

    if filtered:
        cols = st.columns(3)
        for index, item in enumerate(filtered):
            with cols[index % 3]:
                severity = str(item.get("severity", ""))
                st.markdown(
                    f"""
                    <div class="section-card">
                        <span class="{badge_class(severity)}">{severity}</span>
                        <span class="status-badge">{item.get('status', '')}</span>
                        <h3>{item.get('incident_code', '')}</h3>
                        <b>{item.get('title', '')}</b><br><br>
                        Category: {item.get('category', '')}<br>
                        Assigned: {item.get('assigned_to') or 'Unassigned'}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
    else:
        st.info("No matching incidents.")

    show_table(
        filtered,
        ["incident_code", "title", "category", "severity", "status", "assigned_to", "reported_by", "detected_at", "asset_id"],
        "No incidents.",
    )



def human_file_size(size_bytes: int) -> str:
    value = float(size_bytes)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}"
        value /= 1024
    return f"{value:.1f} GB"


def system_status_page() -> None:
    page_header(
        "System Status",
        "Monitor platform health, database connectivity, module data and evidence storage.",
    )

    refresh_col, timestamp_col = st.columns([1, 4])
    with refresh_col:
        if st.button("🔄 Refresh status", width="stretch"):
            st.rerun()

    status_data = api_get("/management/system-status")
    health = api_get("/health")
    root = api_get("/")

    if not status_data:
        st.error("System-status information could not be loaded.")
        return

    counts = status_data.get("record_counts", {})
    storage = status_data.get("storage", {})
    checks = status_data.get("checks", {})

    with timestamp_col:
        st.caption(
            "Last refreshed from server: "
            f"{status_data.get('server_time_utc', 'Unknown')}"
        )

    st.subheader("Service health")
    service_cols = st.columns(4)
    with service_cols[0]:
        kpi(
            "Backend API",
            "Healthy" if status_data.get("backend_status") == "healthy" else "Unavailable",
            "FastAPI service",
        )
    with service_cols[1]:
        kpi(
            "Database",
            "Connected" if status_data.get("database_status") == "connected" else "Disconnected",
            "PostgreSQL connectivity",
        )
    with service_cols[2]:
        kpi(
            "Frontend",
            "Running",
            "Streamlit application",
        )
    with service_cols[3]:
        kpi(
            "Version",
            status_data.get("version", (root or {}).get("version", "Unknown")),
            API_BASE_URL,
        )

    st.subheader("Module records")
    first_row = st.columns(4)
    with first_row[0]:
        kpi("Assets", counts.get("assets", 0), "Registered information assets")
    with first_row[1]:
        kpi("Risks", counts.get("risks", 0), "Risk assessments")
    with first_row[2]:
        kpi("ISO Controls", counts.get("controls", 0), "Control-library records")
    with first_row[3]:
        kpi("Evidence", counts.get("evidence", 0), "Evidence metadata records")

    second_row = st.columns(4)
    with second_row[0]:
        kpi("Audits", counts.get("audits", 0), "Audit engagements")
    with second_row[1]:
        kpi("Findings", counts.get("findings", 0), "Audit findings")
    with second_row[2]:
        kpi(
            "Corrective Actions",
            counts.get("corrective_actions", 0),
            "Remediation actions",
        )
    with second_row[3]:
        kpi("Incidents", counts.get("incidents", 0), "Security incidents")

    st.subheader("Evidence storage")
    storage_cols = st.columns(4)
    with storage_cols[0]:
        kpi(
            "Upload Folder",
            "Available" if storage.get("upload_folder_exists") else "Missing",
            storage.get("upload_folder", "/uploads/evidence"),
        )
    with storage_cols[1]:
        kpi(
            "Folder Access",
            "Accessible" if storage.get("upload_folder_writable") else "Unavailable",
            "Evidence persistence directory",
        )
    with storage_cols[2]:
        kpi(
            "Stored Files",
            storage.get("stored_files", 0),
            "Physical evidence files",
        )
    with storage_cols[3]:
        kpi(
            "Storage Used",
            human_file_size(int(storage.get("storage_bytes", 0))),
            "Current evidence-file storage",
        )

    st.subheader("Operational checks")
    check_labels = {
        "api_responding": "API responding",
        "database_responding": "Database responding",
        "upload_folder_accessible": "Upload folder accessible",
        "iso_controls_loaded": "ISO controls loaded",
        "assets_loaded": "Assets loaded",
        "risks_loaded": "Risks loaded",
    }
    check_columns = st.columns(3)
    for index, (key, label) in enumerate(check_labels.items()):
        passed = bool(checks.get(key))
        with check_columns[index % 3]:
            if passed:
                st.success(f"✓ {label}")
            else:
                st.warning(f"⚠ {label}")

    with st.expander("Technical response"):
        st.json(
            {
                "api_base_url": API_BASE_URL,
                "application": root,
                "health": health,
                "system_status": status_data,
            }
        )


def user_administration_page() -> None:
    page_header(
        "User Administration",
        "Create accounts, assign roles, reset passwords and enable or disable access.",
    )

    users = api_get("/users") or []
    roles = api_get("/roles") or []
    departments = api_get("/departments") or []

    role_by_id = {item["id"]: item["name"] for item in roles}
    department_by_id = {item["id"]: item["name"] for item in departments}
    role_options = {item["name"]: item["id"] for item in roles}
    department_options = {"None": None} | {
        item["name"]: item["id"] for item in departments
    }

    summary_cols = st.columns(3)
    with summary_cols[0]:
        kpi("Total Users", len(users), "Registered accounts")
    with summary_cols[1]:
        kpi(
            "Active Users",
            sum(bool(item.get("is_active")) for item in users),
            "Accounts allowed to sign in",
        )
    with summary_cols[2]:
        kpi(
            "Inactive Users",
            sum(not bool(item.get("is_active")) for item in users),
            "Disabled accounts",
        )

    search_col, role_col, status_col = st.columns([2, 1, 1])
    search = search_col.text_input(
        "Search users",
        placeholder="Name or email",
    )
    selected_role = role_col.selectbox(
        "Role filter",
        ["All"] + sorted(role_options),
    )
    selected_status = status_col.selectbox(
        "Status filter",
        ["All", "Active", "Inactive"],
    )

    filtered_users = users
    if search:
        token = search.lower().strip()
        filtered_users = [
            item
            for item in filtered_users
            if token in str(item.get("full_name", "")).lower()
            or token in str(item.get("email", "")).lower()
        ]

    if selected_role != "All":
        filtered_users = [
            item
            for item in filtered_users
            if role_by_id.get(item.get("role_id")) == selected_role
        ]

    if selected_status != "All":
        expected_active = selected_status == "Active"
        filtered_users = [
            item
            for item in filtered_users
            if bool(item.get("is_active")) == expected_active
        ]

    display_users = [
        {
            "id": item.get("id"),
            "full_name": item.get("full_name"),
            "email": item.get("email"),
            "role": role_by_id.get(item.get("role_id"), "Unknown"),
            "department": department_by_id.get(
                item.get("department_id"),
                "None",
            ),
            "status": "Active" if item.get("is_active") else "Inactive",
            "created_at": item.get("created_at"),
        }
        for item in filtered_users
    ]

    show_table(
        display_users,
        [
            "id",
            "full_name",
            "email",
            "role",
            "department",
            "status",
            "created_at",
        ],
        "No users match the selected filters.",
    )

    with st.expander("➕ Create user"):
        with st.form("create_user_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            full_name = c1.text_input("Full name")
            email = c2.text_input("Email")
            password = c1.text_input(
                "Temporary password",
                type="password",
            )
            role_name = c2.selectbox(
                "Role",
                list(role_options) if role_options else ["No roles"],
            )
            department_name = st.selectbox(
                "Department",
                list(department_options),
            )
            submitted = st.form_submit_button(
                "Create user",
                width="stretch",
            )

        if submitted and role_options:
            created = api_post(
                "/users",
                {
                    "full_name": full_name.strip(),
                    "email": email.strip().lower(),
                    "password": password,
                    "role_id": role_options[role_name],
                    "department_id": department_options[department_name],
                },
            )
            if created:
                st.success("User created.")
                st.rerun()

    if not users:
        return

    user_options = {
        f"{item['full_name']} — {item['email']} (ID {item['id']})": item
        for item in users
    }

    with st.expander("✏️ Edit, enable or disable user"):
        selected_label = st.selectbox(
            "Select user",
            list(user_options),
            key="admin_edit_user",
        )
        selected_user = user_options[selected_label]

        current_role_name = role_by_id.get(
            selected_user.get("role_id"),
            next(iter(role_options), ""),
        )
        current_department_name = department_by_id.get(
            selected_user.get("department_id"),
            "None",
        )

        with st.form("edit_user_form"):
            c1, c2 = st.columns(2)
            edit_full_name = c1.text_input(
                "Full name",
                value=selected_user.get("full_name", ""),
            )
            edit_email = c2.text_input(
                "Email",
                value=selected_user.get("email", ""),
            )
            edit_role = c1.selectbox(
                "Role",
                list(role_options),
                index=list(role_options).index(current_role_name)
                if current_role_name in role_options
                else 0,
            )
            edit_department = c2.selectbox(
                "Department",
                list(department_options),
                index=list(department_options).index(
                    current_department_name
                )
                if current_department_name in department_options
                else 0,
            )
            edit_active = st.checkbox(
                "Account active",
                value=bool(selected_user.get("is_active")),
            )
            save_changes = st.form_submit_button(
                "Save user changes",
                width="stretch",
            )

        if save_changes:
            updated = api_put(
                f"/users/{selected_user['id']}",
                {
                    "full_name": edit_full_name.strip(),
                    "email": edit_email.strip().lower(),
                    "role_id": role_options[edit_role],
                    "department_id": department_options[edit_department],
                    "is_active": edit_active,
                },
            )
            if updated:
                st.success("User updated.")
                st.rerun()

    with st.expander("🔑 Reset user password"):
        selected_password_label = st.selectbox(
            "Select account",
            list(user_options),
            key="admin_reset_user",
        )
        password_user = user_options[selected_password_label]

        with st.form("reset_password_form", clear_on_submit=True):
            new_password = st.text_input(
                "New temporary password",
                type="password",
                help="Minimum 8 characters.",
            )
            confirm_password = st.text_input(
                "Confirm temporary password",
                type="password",
            )
            reset_password = st.form_submit_button(
                "Reset password",
                width="stretch",
            )

        if reset_password:
            if new_password != confirm_password:
                st.error("The password confirmation does not match.")
            elif len(new_password) < 8:
                st.error("The temporary password must contain at least 8 characters.")
            else:
                updated = api_put(
                    f"/users/{password_user['id']}",
                    {"password": new_password},
                )
                if updated:
                    st.success(
                        f"Password reset for {password_user['email']}."
                    )


def audit_trail_page() -> None:
    page_header(
        "Audit Trail",
        "Review authentication events and changes made through the API.",
    )

    logs = api_get("/audit-logs") or []

    if not logs:
        st.info("No audit events have been recorded.")
        return

    df = pd.DataFrame(logs)

    df["created_at"] = pd.to_datetime(
        df["created_at"],
        errors="coerce",
    )

    total_events = len(df)

    successful_logins = len(
        df[df["action"] == "Login successful"]
    )

    failed_logins = len(
        df[df["action"] == "Login failed"]
    )

    today = pd.Timestamp.now().date()

    todays_events = len(
        df[df["created_at"].dt.date == today]
    )

    c1, c2, c3, c4 = st.columns(4)

    with c1:
        kpi(
            "Total Events",
            total_events,
            "All recorded audit events",
        )

    with c2:
        kpi(
            "Successful Logins",
            successful_logins,
            "Successful authentication events",
        )

    with c3:
        kpi(
            "Failed Logins",
            failed_logins,
            "Rejected authentication attempts",
        )

    with c4:
        kpi(
            "Today's Events",
            todays_events,
            "Events recorded today",
        )

    search_col, user_col, action_col, status_col = st.columns(
        [2, 1, 1, 1]
    )

    search = search_col.text_input(
        "Search audit trail",
        placeholder="Email, resource, action or IP address",
    )

    users = [
        "All",
        *sorted(
            df["user_email"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        ),
    ]

    selected_user = user_col.selectbox(
        "User",
        users,
    )

    actions = [
        "All",
        *sorted(
            df["action"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        ),
    ]

    selected_action = action_col.selectbox(
        "Action",
        actions,
    )

    selected_status = status_col.selectbox(
        "Status",
        ["All", "Success", "Failed"],
    )

    filtered = df.copy()

    if search:
        token = search.lower().strip()

        filtered = filtered[
            filtered["user_email"]
            .fillna("")
            .astype(str)
            .str.lower()
            .str.contains(token)
            |
            filtered["resource"]
            .fillna("")
            .astype(str)
            .str.lower()
            .str.contains(token)
            |
            filtered["action"]
            .fillna("")
            .astype(str)
            .str.lower()
            .str.contains(token)
            |
            filtered["ip_address"]
            .fillna("")
            .astype(str)
            .str.lower()
            .str.contains(token)
        ]

    if selected_user != "All":
        filtered = filtered[
            filtered["user_email"] == selected_user
        ]

    if selected_action != "All":
        filtered = filtered[
            filtered["action"] == selected_action
        ]

    if selected_status == "Success":
        filtered = filtered[
            filtered["status_code"] < 400
        ]

    elif selected_status == "Failed":
        filtered = filtered[
            filtered["status_code"] >= 400
        ]

    chart_col, export_col = st.columns([3, 1])

    with chart_col:
        st.subheader("Events by action")

        action_counts = (
            filtered["action"]
            .value_counts()
            .rename_axis("Action")
            .reset_index(name="Events")
        )

        if not action_counts.empty:
            fig = px.bar(
                action_counts,
                x="Action",
                y="Events",
                text_auto=True,
            )

            fig.update_layout(
                height=320,
                margin=dict(
                    l=10,
                    r=10,
                    t=20,
                    b=10,
                ),
            )

            st.plotly_chart(
                fig,
                width="stretch",
            )

    display_df = filtered.copy()

    display_df["Status"] = display_df[
        "status_code"
    ].apply(
        lambda value: (
            "🟢 Success"
            if int(value) < 400
            else "🔴 Failed"
        )
    )

    display_df = display_df.rename(
        columns={
            "created_at": "Time",
            "user_email": "User",
            "action": "Action",
            "resource": "Resource",
            "method": "Method",
            "ip_address": "IP",
        }
    )

    display_df = display_df[
        [
            "Time",
            "User",
            "Action",
            "Resource",
            "Method",
            "Status",
            "IP",
        ]
    ]

    display_df = display_df.sort_values(
        by="Time",
        ascending=False,
    )

    with export_col:
        csv_data = display_df.to_csv(
            index=False
        ).encode("utf-8")

        st.download_button(
            "⬇ Export CSV",
            data=csv_data,
            file_name="audit_trail.csv",
            mime="text/csv",
            width="stretch",
        )

        st.metric(
            "Filtered Events",
            len(display_df),
        )

    st.subheader("Audit events")

    if display_df.empty:
        st.info("No audit events match the selected filters.")
    else:
        st.dataframe(
            display_df,
            width="stretch",
            hide_index=True,
        )



def reports_center_page() -> None:
    page_header(
        "Reports Center",
        "Generate management-ready PDF reports and consolidated Excel workbooks.",
    )

    current_user = st.session_state.get("current_user") or {}
    generated_by = current_user.get("full_name") or current_user.get("email") or "Authorized user"

    management = api_get("/management/dashboard") or {}
    compliance = api_get("/compliance/dashboard") or {}
    compliance_summary = api_get("/compliance/summary") or {}
    risks = api_get("/risks") or []
    controls = api_get("/controls") or []
    assessments = api_get("/compliance-assessments") or []
    audits = api_get("/audits") or []
    findings = api_get("/audit-findings") or []
    actions = api_get("/corrective-actions") or []
    incidents = api_get("/incidents") or []
    evidence = api_get("/evidence") or []

    summary_cols = st.columns(4)
    with summary_cols[0]:
        kpi("Available Reports", 6, "PDF report templates")
    with summary_cols[1]:
        kpi("Risks", len(risks), "Risk-register records")
    with summary_cols[2]:
        kpi("Controls", len(controls), "ISO control records")
    with summary_cols[3]:
        kpi("Generated By", generated_by, "Authenticated report owner")

    management_summary = [
        ("Total assets", management.get("total_assets", 0)),
        ("Total risks", management.get("total_risks", len(risks))),
        ("Critical risks", management.get("critical_risks", 0)),
        ("Compliance percentage", f"{management.get('compliance_percentage', 0):.2f}%"),
        ("Open incidents", management.get("open_incidents", 0)),
        ("Open audit findings", management.get("open_audit_findings", 0)),
        ("Overdue corrective actions", management.get("overdue_corrective_actions", 0)),
    ]

    reports = [
        {
            "title": "Executive Management Report",
            "description": "Executive overview of risk, compliance, incidents, findings and remediation.",
            "file_name": "cyber_x_force_executive_management_report.pdf",
            "summary": management_summary,
            "sections": [
                ("Risk Register", risks),
                ("ISO Controls", controls),
                ("Audit Findings", findings),
                ("Corrective Actions", actions),
                ("Incidents", incidents),
            ],
        },
        {
            "title": "Risk Register Report",
            "description": "Detailed risk register with inherent and residual risk information.",
            "file_name": "cyber_x_force_risk_register_report.pdf",
            "summary": [
                ("Total risks", len(risks)),
                ("Critical risks", sum(item.get("inherent_level") == "Critical" for item in risks)),
                ("High risks", sum(item.get("inherent_level") == "High" for item in risks)),
                ("Open risks", sum(item.get("status") != "Closed" for item in risks)),
            ],
            "sections": [("Risk Register", risks)],
        },
        {
            "title": "ISO 27001 Compliance Report",
            "description": "Control implementation, compliance assessments and overdue-control status.",
            "file_name": "cyber_x_force_iso27001_compliance_report.pdf",
            "summary": [
                ("Total controls", compliance_summary.get("total_controls", len(controls))),
                ("Implemented controls", compliance_summary.get("implemented", 0)),
                ("Compliance percentage", f"{compliance_summary.get('overall_compliance_percentage', 0):.2f}%"),
                ("Average progress", f"{compliance_summary.get('average_implementation_percentage', 0):.2f}%"),
                ("Assessments", len(assessments)),
            ],
            "sections": [
                ("ISO Controls", controls),
                ("Compliance Assessments", assessments),
            ],
        },
        {
            "title": "Audit and Findings Report",
            "description": "Audit programme, audit findings, related evidence and remediation status.",
            "file_name": "cyber_x_force_audit_findings_report.pdf",
            "summary": [
                ("Audits", len(audits)),
                ("Audit findings", len(findings)),
                ("Open findings", sum(item.get("status") not in {"Closed", "Verified"} for item in findings)),
                ("Evidence records", len(evidence)),
                ("Corrective actions", len(actions)),
            ],
            "sections": [
                ("Audits", audits),
                ("Audit Findings", findings),
                ("Corrective Actions", actions),
                ("Evidence Register", evidence),
            ],
        },
        {
            "title": "Incident Summary Report",
            "description": "Security incident register with severity, status and ownership information.",
            "file_name": "cyber_x_force_incident_summary_report.pdf",
            "summary": [
                ("Total incidents", len(incidents)),
                ("Open incidents", sum(item.get("status") != "Closed" for item in incidents)),
                ("Critical incidents", sum(item.get("severity") == "Critical" for item in incidents)),
                ("High incidents", sum(item.get("severity") == "High" for item in incidents)),
            ],
            "sections": [("Incidents", incidents)],
        },
        {
            "title": "Corrective Action Report",
            "description": "Remediation actions, ownership, due dates and completion status.",
            "file_name": "cyber_x_force_corrective_action_report.pdf",
            "summary": [
                ("Total actions", len(actions)),
                ("Open actions", sum(item.get("status") == "Open" for item in actions)),
                ("In-progress actions", sum(item.get("status") == "In Progress" for item in actions)),
                ("Completed or verified", sum(item.get("status") in {"Completed", "Verified"} for item in actions)),
            ],
            "sections": [
                ("Corrective Actions", actions),
                ("Audit Findings", findings),
            ],
        },
    ]

    st.subheader("PDF reports")
    report_columns = st.columns(2)
    for index, report in enumerate(reports):
        with report_columns[index % 2]:
            st.markdown(
                f"""
                <div class="section-card">
                    <h3>{report['title']}</h3>
                    <span style="opacity:.72">{report['description']}</span>
                </div>
                """,
                unsafe_allow_html=True,
            )
            pdf_data = build_pdf_report(
                report["title"],
                report["description"],
                report["summary"],
                report["sections"],
                generated_by,
            )
            st.download_button(
                f"Download {report['title']}",
                data=pdf_data,
                file_name=report["file_name"],
                mime="application/pdf",
                key=f"pdf-report-{index}",
                width="stretch",
            )

    st.subheader("Management workbook")
    excel_download_button(
        "Download consolidated Excel workbook",
        "cyber_x_force_management_reporting.xlsx",
        {
            "Management Summary": [management],
            "Compliance Summary": [compliance_summary],
            "Risks": risks,
            "ISO Controls": controls,
            "Assessments": assessments,
            "Audits": audits,
            "Audit Findings": findings,
            "Corrective Actions": actions,
            "Incidents": incidents,
            "Evidence": evidence,
        },
        "Cyber_X_Force Management Reporting Workbook",
        "reports-center-workbook",
    )

    with st.expander("Report metadata"):
        st.json(
            {
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "generated_by": generated_by,
                "application": "Cyber_X_Force",
                "framework": "ISO/IEC 27001",
                "api_base_url": API_BASE_URL,
            }
        )


PAGES = {
    "🏠 Dashboard": dashboard_page,
    "🗂️ Assets": assets_page,
    "⚠️ Risk Register": risks_page,
    "🛡️ ISO Controls": controls_page,
    "✅ Compliance": compliance_page,
    "📁 Evidence": evidence_page,
    "🔎 Audits and Findings": audits_page,
    "🧰 Corrective Actions": actions_page,
    "🚨 Incidents": incidents_page,
    "⚙️ System Status": system_status_page,
    "📊 Reports Center": reports_center_page,
}


ROLE_PAGES = {
    "Administrator": list(PAGES) + ["👥 User Administration", "📜 Audit Trail"],
    "Risk Manager": ["🏠 Dashboard", "🗂️ Assets", "⚠️ Risk Register", "🛡️ ISO Controls", "✅ Compliance", "📁 Evidence", "⚙️ System Status"],
    "Asset Owner": ["🏠 Dashboard", "🗂️ Assets", "⚠️ Risk Register", "🛡️ ISO Controls", "📁 Evidence"],
    "Internal Auditor": ["🏠 Dashboard", "🛡️ ISO Controls", "✅ Compliance", "📁 Evidence", "🔎 Audits and Findings", "🧰 Corrective Actions", "📜 Audit Trail", "📊 Reports Center"],
    "Incident Manager": ["🏠 Dashboard", "🗂️ Assets", "⚠️ Risk Register", "🧰 Corrective Actions", "🚨 Incidents", "📁 Evidence"],
    "Executive Viewer": ["🏠 Dashboard", "✅ Compliance", "⚙️ System Status", "📊 Reports Center"],
}
EXTRA_PAGES = {"👥 User Administration": user_administration_page, "📜 Audit Trail": audit_trail_page}
ALL_PAGES = PAGES | EXTRA_PAGES

if "access_token" not in st.session_state:
    st.session_state.access_token = None
if "current_user" not in st.session_state:
    st.session_state.current_user = None

def login_page() -> None:
    st.markdown("<br><br>", unsafe_allow_html=True)
    left, center, right = st.columns([1, 1.2, 1])
    with center:
        page_header("🛡️ Cyber_X_Force Login", "Sign in to the ISO/IEC 27001 GRC platform.")
        with st.form("login_form"):
            email = st.text_input("Email", value="admin@cyberxforce.com")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login", width="stretch")
        if submitted:
            result = api_post("/auth/login", {"email": email, "password": password})
            if result:
                st.session_state.access_token = result["access_token"]
                st.session_state.current_user = result["user"]
                st.rerun()
        st.caption("Demo administrator: admin@cyberxforce.com / Admin@12345")

if not st.session_state.access_token:
    login_page()
    st.stop()

if not st.session_state.current_user:
    st.session_state.current_user = api_get("/auth/me")
if not st.session_state.current_user:
    st.session_state.access_token = None
    st.stop()

current_user = st.session_state.current_user
role_name = current_user.get("role_name", "Executive Viewer")
available_pages = [page for page in ROLE_PAGES.get(role_name, ["🏠 Dashboard"]) if page in ALL_PAGES]

with st.sidebar:
    st.markdown("## 🛡️ Cyber_X_Force")
    st.caption("ISO/IEC 27001 GRC Platform")
    st.success(f"{current_user.get('full_name')}\n\n{role_name}")
    selected_page = st.radio("Navigation", available_pages, label_visibility="collapsed")
    st.divider()
    health = api_get("/health")
    if health and health.get("status") == "healthy":
        st.success("Backend connected")
    else:
        st.error("Backend unavailable")
    if st.button("Logout", width="stretch"):
        st.session_state.access_token = None
        st.session_state.current_user = None
        st.rerun()
    st.caption(API_BASE_URL)

ALL_PAGES[selected_page]()
